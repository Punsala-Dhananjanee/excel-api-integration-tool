import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .logger import get_logger

logger = get_logger(__name__)

# Style constants
HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
DATA_FONT     = Font(name="Arial", size=10)
ALT_ROW_FILL  = PatternFill("solid", start_color="EBF3FB")
SUMMARY_FILL  = PatternFill("solid", start_color="2E75B6")
SUMMARY_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=12)
THIN_BORDER   = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


class ExcelWriter:
    def __init__(self, config: dict):
        self.cfg = config.get("excel", {})
        self.output_dir = self.cfg.get("output_dir", "data/output")
        os.makedirs(self.output_dir, exist_ok=True)

    def _resolve_path(self) -> str:
        filename = self.cfg.get("filename")
        if filename:
            return os.path.join(self.output_dir, filename)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(self.output_dir, f"api_data_{ts}.xlsx")

    def write(self, data: dict) -> str:
        """Write all sheet data to Excel. Returns output path."""
        path = self._resolve_path()
        overwrite = self.cfg.get("overwrite", True)

        if os.path.exists(path) and not overwrite:
            logger.info(f"File exists and overwrite=False, skipping: {path}")
            return path

        wb = Workbook()
        wb.remove(wb.active)  # remove default blank sheet

        if self.cfg.get("add_summary_sheet", True):
            self._write_summary(wb, data)

        for sheet_name, rows in data.items():
            self._write_sheet(wb, sheet_name, rows)

        wb.save(path)
        logger.info(f"Workbook saved: {path}")
        return path

    def _write_summary(self, wb: Workbook, data: dict):
        ws = wb.create_sheet("Summary", 0)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D1")
        ws["A1"] = "📊 API Data Integration — Summary"
        ws["A1"].fill = SUMMARY_FILL
        ws["A1"].font = SUMMARY_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        ws["A3"] = "Generated At"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A4"] = "Total Sheets"
        ws["B4"] = len(data)

        ws["A6"] = "Sheet"
        ws["B6"] = "Row Count"
        for cell in [ws["A6"], ws["B6"]]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for i, (name, rows) in enumerate(data.items(), start=7):
            ws[f"A{i}"] = name
            ws[f"B{i}"] = len(rows)
            if i % 2 == 0:
                for col in ["A", "B"]:
                    ws[f"{col}{i}"].fill = ALT_ROW_FILL

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def _write_sheet(self, wb: Workbook, sheet_name: str, rows: list):
        if not rows:
            logger.warning(f"No rows for sheet '{sheet_name}', skipping.")
            return

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars
        ws.sheet_view.showGridLines = False

        headers = list(rows[0].keys())

        # Write header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 22

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
            for col_idx, header in enumerate(headers, start=1):
                value = row.get(header)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if fill:
                    cell.fill = fill

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row in rows:
                val = row.get(header, "")
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # Auto-filter
        if self.cfg.get("auto_filter", True) and rows:
            ws.auto_filter.ref = ws.dimensions

        # Freeze panes (freeze header row)
        if self.cfg.get("freeze_panes", True):
            ws.freeze_panes = "A2"

        logger.debug(f"Sheet '{sheet_name}' written ({len(rows)} rows, {len(headers)} cols)")
